import pandas as pd
from typing import List, Dict
from datetime import datetime
from ..models.models import VenueReservation, DeviceReservation, PrinterReservation
import io
import bcrypt
from io import BytesIO
from openpyxl.utils import get_column_letter

def read_users_excel(file) -> List[Dict]:
    df = pd.read_excel(file)
    users = []
    
    for _, row in df.iterrows():
        username = str(row["学号/工号"])
        password = str(row["身份证号后6位"])
        user = {
            "username": username,
            "name": str(row["姓名"]),
            "department": str(row["学院"]),
            "password": bcrypt.hashpw(
                password.encode('utf-8'), 
                bcrypt.gensalt()
            ).decode('utf-8'),
            "role": "student" if len(username) > 6 else "teacher"
        }
        users.append(user)
    
    return users

def format_reservation_status(status: str) -> str:
    """格式化预约状态为中文"""
    status_map = {
        "pending": "待审批",
        "approved": "已通过",
        "rejected": "已拒绝",
        "returned": "已归还"
    }
    return status_map.get(status, status)

def format_device_name(name: str) -> str:
    """格式化设备名称为中文"""
    device_map = {
        "electric_screwdriver": "电动螺丝刀",
        "multimeter": "万用表",
        "printer_1": "3D打印机1号",
        "printer_2": "3D打印机2号",
        "printer_3": "3D打印机3号"
    }
    return device_map.get(name, name)

def export_reservations_excel(
    venue_reservations: List[VenueReservation],
    device_reservations: List[DeviceReservation],
    printer_reservations: List[PrinterReservation]
) -> bytes:
    """导出预约记录到Excel"""
    try:
        # 准备数据
        venue_data = []
        for res in venue_reservations:
            # 处理设备需求
            devices_needed = []
            if res.devices_needed:
                if res.devices_needed.get('screen'): devices_needed.append('大屏')
                if res.devices_needed.get('laptop'): devices_needed.append('笔记本')
                if res.devices_needed.get('mic_handheld'): devices_needed.append('手持麦')
                if res.devices_needed.get('mic_gooseneck'): devices_needed.append('鹅颈麦')
                if res.devices_needed.get('projector'): devices_needed.append('投屏器')

            venue_data.append({
                "预约类型": "场地预约",
                "场地类型": str(res.venue_type),
                "预约日期": res.reservation_date.strftime("%Y-%m-%d"),
                "时间段": str(res.business_time),
                "用途": str(res.purpose),
                "设备需求": '、'.join(devices_needed) if devices_needed else '无',
                "预约人": f"{str(res.user.name)}({str(res.user.username)})",
                "所属部门": str(res.user.department),
                "状态": str(format_reservation_status(res.status))
            })
        
        device_data = []
        for res in device_reservations:
            device_data.append({
                "预约类型": "设备预约",
                "设备名称": str(format_device_name(res.device_name)),
                "借用时间": res.borrow_time.strftime("%Y-%m-%d %H:%M"),
                "预计归还时间": res.return_time.strftime("%Y-%m-%d %H:%M"),
                "实际归还时间": res.actual_return_time.strftime("%Y-%m-%d %H:%M") if res.actual_return_time else "未归还",
                "借用原因": str(res.reason),
                "预约人": f"{str(res.user.name)}({str(res.user.username)})",
                "所属部门": str(res.user.department),
                "状态": str(format_reservation_status(res.status))
            })
        
        printer_data = []
        for res in printer_reservations:
            printer_data.append({
                "预约类型": "3D打印机预约",
                "打印机": str(format_device_name(res.printer_name)),
                "预约日期": res.reservation_date.strftime("%Y-%m-%d"),
                "打印时间": res.print_time.strftime("%H:%M"),
                "预约人": f"{str(res.user.name)}({str(res.user.username)})",
                "所属部门": str(res.user.department),
                "状态": str(format_reservation_status(res.status))
            })

        # 创建DataFrame
        dfs = []
        if venue_data:
            venue_df = pd.DataFrame(venue_data)
            # 设置场地预约的列顺序
            venue_columns = [
                "预约类型", "场地类型", "预约日期", "时间段", 
                "用途", "设备需求", "预约人", "所属部门", "状态"
            ]
            venue_df = venue_df.reindex(columns=venue_columns)
            dfs.append(("场地预约", venue_df))
            
        if device_data:
            device_df = pd.DataFrame(device_data)
            # 设置设备预约的列顺序
            device_columns = [
                "预约类型", "设备名称", "借用时间", "预计归还时间", 
                "实际归还时间", "借用原因", "预约人", "所属部门", "状态"
            ]
            device_df = device_df.reindex(columns=device_columns)
            dfs.append(("设备预约", device_df))
            
        if printer_data:
            printer_df = pd.DataFrame(printer_data)
            # 设置打印机预约的列顺序
            printer_columns = [
                "预约类型", "打印机", "预约日期", "打印时间",
                "预约人", "所属部门", "状态"
            ]
            printer_df = printer_df.reindex(columns=printer_columns)
            dfs.append(("3D打印机预约", printer_df))

        # 使用openpyxl直接写入
        from openpyxl import Workbook
        wb = Workbook()
        
        # 删除默认的sheet
        wb.remove(wb.active)
        
        # 写入每个sheet
        for sheet_name, df in dfs:
            ws = wb.create_sheet(sheet_name)
            # 写入表头
            for col, header in enumerate(df.columns, 1):
                ws.cell(row=1, column=col, value=header)
            # 写入数据
            for row_idx, row in enumerate(df.values, 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 调整列宽
            for idx, col in enumerate(df.columns, 1):
                max_length = max(
                    df[col].astype(str).apply(len).max(),
                    len(col)
                )
                ws.column_dimensions[get_column_letter(idx)].width = max_length + 2

        # 保存到BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()

    except Exception as e:
        print(f"Error in export_reservations_excel: {str(e)}")
        raise 